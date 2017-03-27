"""
Given input article DOI, inserts article metadata (Publication type & MeSH headings)
into excel sheets

currently set to process batch of 200 DOIs
python efetch_complete2.py

"""
import re
from Bio import Entrez
from openpyxl import load_workbook


def search(doi):
    Entrez.email = 'dclynch123@gmail.com'
    handle = Entrez.esearch(db='pubmed', retmode='xml', term=doi)
    results = Entrez.read(handle)
    return results

def fetch_details(id_list):
    Entrez.email = 'dclynch123@gmail.com'
    handle = Entrez.efetch(db='pubmed', retmode='xml', id=id_list)
    results = Entrez.read(handle)
    return results


if __name__ == '__main__':
    wb = load_workbook(filename = 'backlog_DOIs.xlsx')
    mesh_sheet = wb.get_sheet_by_name('mesh')
    pub_sheet = wb.get_sheet_by_name('pubtype')

    for rowNum in range(2, 202):
        
        # print progress
        print "Starting row number {}".format(rowNum)

        # initialize counters
        mesh_data = 0
        no_mesh_data = 0
        pub_data = 0
        no_pub_data = 0
        
        # lookup doi
        doi = mesh_sheet.cell(row=rowNum, column=2).value
        results = search(doi)
        try:
            id_list = results['IdList'][0]
            res = fetch_details(id_list)
        except:
            mesh_sheet.cell(row=rowNum, column=3).value = "No article data"

        # if pub type exists append to pubtype list and assign to cell
        try:
            pubtype_list = res['PubmedArticle'][0]['MedlineCitation']['Article']['PublicationTypeList']
            pubtype_str = str(pubtype_list)
            pub_sheet.cell(row=rowNum, column=3).value = pubtype_str
            ++pub_data
        except:
            pub_sheet.cell(row=rowNum, column=3).value = "No pubtype data"
            ++no_pub_data

        # initialize mesh list
        mesh_list = []
        # if mesh exists append each mesh heading to mesh list and assign list to cell in spreadsheet
        try:
            myList = res['PubmedArticle'][0]['MedlineCitation']['MeshHeadingList']
            #for attr, value in enumerate(myList):
            #    mesh_list.append(attr, value)
            mesh_str = str(myList)
            mesh_sheet.cell(row=rowNum, column=4).value = mesh_str
            ++mesh_data
        except:
            mesh_sheet.cell(row=rowNum, column=4).value = "No MeSH available"
            ++no_mesh_data

    wb.save('backlog_DOIs.xlsx')
    print "Process completed"
    print ""
    print "MeSH Data was available for {} records".format(mesh_data)
    print "Pubtype data was available for {} records".format(pub_data)
    print ""
    print "Totals: "
    print "Mesh: {}; No mesh: {}; Pubt type: {}; No pub type: {}".format(mesh_data, no_mesh_data, pub_data, no_pub_data)





    pub = res['PubmedArticle'][0]['MedlineCitation']['Article']['PublicationTypeList']
    print ""

    uid = re.compile('[A-Z]\d{6}')
    pubIDs = uid.findall(str(pub))

    for item in pubIDs:
        print "{} | {} ".format(query, item)

    print ""

    # if mesh exists, assign mesh data to first sheet, each mesh heading

    major = re.compile(r"\'[A-Z]\'")

    if (res['PubmedArticle'][0]['MedlineCitation']['MeshHeadingList']):
        print "MeSH Headings:"
        myList = res['PubmedArticle'][0]['MedlineCitation']['MeshHeadingList']
        me = re.compile(r"\'(.+?)\'")

        print ""    
        for value in enumerate(myList):
            print "{} | {} | {}".format(query, uid.findall(str(value)), major.findall(str(value)))
        print ""
    else:
        print "No MeSH available"